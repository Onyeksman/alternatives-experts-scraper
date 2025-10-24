import asyncio
import logging
from typing import Optional, List, Dict
from urllib.parse import urljoin
from datetime import datetime

from playwright.async_api import async_playwright, TimeoutError as pwTimeout
from bs4 import BeautifulSoup
import pandas as pd
from tenacity import retry, stop_after_attempt, wait_exponential, RetryError
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ----------------
START_URL = "https://www.alternatives.org.uk/experts"
BASE = "https://www.alternatives.org.uk"
PAGE_TIMEOUT_MS = 5000
RETRY_ATTEMPTS = 3
OUTPUT_FILE = "speakers.xlsx"

# ---------------- LOGGING ----------------
logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
logger = logging.getLogger("alternatives-scraper")


@retry(stop=stop_after_attempt(RETRY_ATTEMPTS), wait=wait_exponential(multiplier=1, min=1, max=10), reraise=True)
async def fetch_detail_html(context, url: str) -> Optional[str]:
    """Fetch detail page HTML, retrying on transient failures."""
    page = await context.new_page()
    try:
        await page.goto(url, timeout=PAGE_TIMEOUT_MS, wait_until="domcontentloaded")
        await page.wait_for_selector("div.field-content", timeout=PAGE_TIMEOUT_MS)
        await asyncio.sleep(0.2)
        return await page.content()
    finally:
        try:
            await page.close()
        except Exception:
            pass


def parse_main_cards(html: str) -> List[Dict]:
    """Parse main page and return list of speakers in exact on-site order."""
    soup = BeautifulSoup(html, "html.parser")
    nodes = soup.select("div.views-row")
    results = []
    for node in nodes:
        a = node.select_one("h3 a")
        name = a.get_text(strip=True) if a else ""
        href = a.get("href") if a else None
        detail_url = urljoin(BASE, href) if href else None

        first_tag = last_tag = ""
        ul = node.find("ul")
        if ul:
            lis = ul.find_all("li")
            if lis:
                first_tag = lis[0].get_text(strip=True) if lis[0] else ""
                last_tag = lis[-1].get_text(strip=True) if lis[-1] else ""

        results.append({
            "Name": name.strip(),
            "First Tag": first_tag.strip(),
            "Last Tag": last_tag.strip(),
            "Detail URL": detail_url
        })
    return results


def parse_about_from_html(html: Optional[str]) -> str:
    """Extract About section text."""
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    field_blocks = soup.select("div.field-content")
    parts = []
    for block in field_blocks:
        ps = block.select("p")
        if ps:
            for p in ps:
                text = p.get_text(" ", strip=True)
                if text:
                    parts.append(text)
        else:
            text = block.get_text(" ", strip=True)
            if text and len(text) > 30:
                parts.append(text)
    return "\n\n".join(parts).strip()


async def main():
    speakers: List[Dict] = []
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5993.90 Safari/537.36")
        )
        page = await context.new_page()

        # --- Load main page ---
        try:
            logger.info("Loading main page...")
            await page.goto(START_URL, timeout=30000, wait_until="networkidle")
            await asyncio.sleep(0.5)
            main_html = await page.content()
        except pwTimeout:
            logger.warning("Timeout, retrying with domcontentloaded...")
            await page.goto(START_URL, timeout=30000, wait_until="domcontentloaded")
            main_html = await page.content()
        finally:
            await page.close()

        # --- Parse speaker cards in order ---
        records = parse_main_cards(main_html)
        logger.info(f"Found {len(records)} speakers on main page.")

        # --- Sequentially fetch details to preserve order ---
        for idx, rec in enumerate(records, start=1):
            detail_url = rec.get("Detail URL")
            about_text = ""
            if detail_url:
                try:
                    html = await fetch_detail_html(context, detail_url)
                    about_text = parse_about_from_html(html)
                except (RetryError, pwTimeout, Exception):
                    about_text = ""

            speakers.append({
                "Name": rec.get("Name", "").strip(),
                "First Tag": rec.get("First Tag", "").strip(),
                "Last Tag": rec.get("Last Tag", "").strip(),
                "About the Speaker": about_text.strip()
            })

            if idx % 10 == 0:
                logger.info(f"Processed {idx} speakers...")

        await context.close()
        await browser.close()

    # --- Clean and prepare data ---
    df = pd.DataFrame(speakers, columns=["Name", "First Tag", "Last Tag", "About the Speaker"])
    df.replace(["", None], "N/A", inplace=True)
    df.drop_duplicates(inplace=True)
    df.reset_index(drop=True, inplace=True)

    # --- Save and format Excel ---
    df.to_excel(OUTPUT_FILE, index=False)
    style_excel(OUTPUT_FILE)
    logger.info(f"✅ Saved {len(df)} speakers to {OUTPUT_FILE}")


def style_excel(file_path: str):
    """Apply full professional Excel styling."""
    wb = load_workbook(file_path)
    ws = wb.active

    # --- Define styles ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    grey_font = Font(color="808080", italic=True)
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    border_side = Side(border_style="medium", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # --- Header row ---
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Data rows ---
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if idx % 2 == 0:  # Alternate shading
            for cell in row:
                cell.fill = alt_fill
        for cell in row:
            if str(cell.value).strip() == "N/A":
                cell.font = grey_font
            cell.border = border

    # --- Autofit column widths ---
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 60)

    # --- Freeze header & enable filter ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Metadata note at bottom ---
    last_row = ws.max_row + 2
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=last_row, column=1, value="📊 Sourced from (https://www.alternatives.org.uk/experts)")
    ws.cell(row=last_row + 1, column=1, value=f"Generated on: {timestamp}")
    ws.cell(row=last_row, column=1).font = grey_font
    ws.cell(row=last_row + 1, column=1).font = grey_font

    wb.save(file_path)
    wb.close()


if __name__ == "__main__":
    asyncio.run(main())
